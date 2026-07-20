"""Export gold-standard classification comparison to XLSX.

Compares ``data/gold_standard/gold_standard_mcq_100.jsonl`` (frontier
reference) against:

* **embedding** — classified rows under ``data/processed/`` excluding
  ``qwen``, ``qwen_1``, ``qwen_combined`` and comparison/gold files.
* **nli** — classified rows under ``data/processed/qwen_combined``,
  ``data/processed/qwen``, and ``data/processed/qwen_1`` (or the matching
  trees under ``_archive/`` if those were moved).

Rows are matched by question / primary text fields. The workbook includes
question, answers, gold reasons, predictions, per-row match flags, and a
summary sheet with similarity percentages.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from classify_zero_shot import text_for_classification
PROCESSED = ROOT / "data" / "processed"
GOLD = ROOT / "data" / "gold_standard"
SCHEMA = ROOT / "schema"
DEFAULT_GOLD = GOLD / "gold_standard_mcq_100.jsonl"
DEFAULT_OUTPUT = GOLD / "gold_standard_comparison.xlsx"
DEFAULT_OUTPUT_CSV = GOLD / "gold_standard_mcq_comparison.csv"
DEFAULT_COARSE_TAXONOMY = SCHEMA / "taxonomy_coarse.json"

SKIP_FILENAMES = frozenset({
    "gold_standard_mcq_100.jsonl",
    "gold_standard_100.jsonl",
    "gold_standard_classified_frontier.jsonl",
    "classified_with_taxonomies.jsonl",
    "compare_qwen.jsonl",
    "compare_llm2clip.jsonl",
})

EMBEDDING_EXCLUDE_DIRS = frozenset({"qwen", "qwen_1", "qwen_combined"})

NLI_ROOTS = (
    PROCESSED / "qwen_1",
    PROCESSED / "qwen",
    PROCESSED / "qwen_combined",
)

TAXONOMY_KEYS = ("tasks", "risks", "difficulty", "cognitive", "tiers")

_QUESTION_WS_RE = re.compile(r"\s+")


@dataclass
class CorpusIndex:
    by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_question: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_text: dict[str, dict[str, Any]] = field(default_factory=dict)


@dataclass(frozen=True)
class CoarseTaxonomy:
    by_fine_id: dict[str, tuple[str, str]]
    by_coarse_id: dict[str, str]


def load_coarse_taxonomy(path: Path) -> CoarseTaxonomy:
    with path.open(encoding="utf-8") as handle:
        doc = json.load(handle)
    by_fine_id: dict[str, tuple[str, str]] = {}
    by_coarse_id: dict[str, str] = {}
    for row in doc.get("labels", []):
        if not isinstance(row, dict):
            continue
        coarse_id = row.get("id")
        coarse_name = row.get("name")
        members = row.get("members")
        if not isinstance(coarse_id, str) or not isinstance(coarse_name, str):
            continue
        if not isinstance(members, list):
            continue
        by_coarse_id[coarse_id] = coarse_name
        for member in members:
            if isinstance(member, str):
                by_fine_id[member] = (coarse_id, coarse_name)
    if not by_fine_id:
        raise ValueError(f"{path}: no fine-to-coarse mappings found")
    return CoarseTaxonomy(by_fine_id=by_fine_id, by_coarse_id=by_coarse_id)


def coarse_group(fine_label_id: str, coarse: CoarseTaxonomy) -> tuple[str, str]:
    if not fine_label_id:
        return "", ""
    hit = coarse.by_fine_id.get(fine_label_id)
    if hit is None:
        return "", ""
    return hit


def iter_jsonl(path: Path) -> Any:
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                yield json.loads(line)


def is_skipped_jsonl(path: Path) -> bool:
    if path.name in SKIP_FILENAMES:
        return True
    if "top_n" in path.name:
        return True
    return False


def classification_text(record: dict[str, Any]) -> str:
    payload = record.get("payload")
    if isinstance(payload, dict):
        return text_for_classification(payload)
    return text_for_classification(record)


def normalize_question_key(text: str) -> str:
    return _QUESTION_WS_RE.sub(" ", text.strip())


def question_key_from_record(record: dict[str, Any]) -> str | None:
    payload = record.get("payload")
    if isinstance(payload, dict):
        question = payload.get("question")
        if isinstance(question, str) and question.strip():
            return normalize_question_key(question)
    for key in ("question", "text"):
        value = record.get(key)
        if isinstance(value, str) and value.strip():
            return normalize_question_key(value)
    return None


def match_keys_from_record(record: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        stripped = value.strip()
        if stripped and stripped not in seen:
            seen.add(stripped)
            keys.append(stripped)

    text = record.get("text")
    if isinstance(text, str):
        add(text)

    add(classification_text(record))
    return keys


def format_answers(payload: dict[str, Any]) -> str:
    if not payload:
        return ""

    answers = payload.get("answers")
    if isinstance(answers, dict) and answers:
        return "\n".join(f"{key}) {value}" for key, value in answers.items())

    choices = payload.get("choices")
    if isinstance(choices, dict) and choices:
        return "\n".join(f"{key}) {value}" for key, value in choices.items())

    if isinstance(choices, list) and choices:
        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        return "\n".join(
            f"{letters[i] if i < len(letters) else i}) {choice}"
            for i, choice in enumerate(choices)
        )

    options = payload.get("options")
    if isinstance(options, list) and options:
        return "\n".join(f"- {opt}" for opt in options)

    option_lines: list[str] = []
    for letter in "abcdefghijklmnopqrstuvwxyz":
        key = f"option_{letter}"
        if key in payload and payload[key] is not None:
            option_lines.append(f"{letter.upper()}) {payload[key]}")
    if option_lines:
        return "\n".join(option_lines)

    return ""


def format_solution(record: dict[str, Any], payload: dict[str, Any]) -> str:
    evaluation = record.get("evaluation")
    if isinstance(evaluation, dict):
        answer = evaluation.get("correct_answer")
        if answer is not None and answer != "":
            return str(answer)
    for key in ("solution", "answer", "label", "gold_label"):
        value = payload.get(key)
        if value is not None and value != "":
            return str(value)
    return ""


def classification_topic(classification: dict[str, Any] | None) -> tuple[str, str, str]:
    if not isinstance(classification, dict):
        return "", "", ""
    label = str(classification.get("predicted_label") or "")
    name = str(classification.get("predicted_label_name") or "")
    reason = str(classification.get("reason") or "")
    return label, name, reason


def taxonomy_field(classification: dict[str, Any] | None, key: str) -> tuple[str, str, str]:
    if not isinstance(classification, dict):
        return "", "", ""
    taxonomies = classification.get("taxonomies")
    if not isinstance(taxonomies, dict):
        return "", "", ""
    row = taxonomies.get(key)
    if not isinstance(row, dict):
        return "", "", ""
    label = str(row.get("predicted_label") or "")
    name = str(row.get("predicted_label_name") or label)
    reason = str(row.get("reason") or "")
    return label, name, reason


def build_index(roots: list[Path], *, exclude_dirs: set[str] | None = None) -> CorpusIndex:
    index = CorpusIndex()
    for root in roots:
        if not root.is_dir():
            continue
        for path in sorted(root.rglob("*.jsonl")):
            if is_skipped_jsonl(path):
                continue
            if exclude_dirs is not None:
                try:
                    rel = path.relative_to(PROCESSED)
                    if rel.parts and rel.parts[0] in exclude_dirs:
                        continue
                except ValueError:
                    pass
            try:
                records = list(iter_jsonl(path))
            except json.JSONDecodeError:
                continue
            for record in records:
                record_id = record.get("id")
                if isinstance(record_id, str) and record_id:
                    index.by_id[record_id] = record
                question_key = question_key_from_record(record)
                if question_key:
                    index.by_question[question_key] = record
                for key in match_keys_from_record(record):
                    index.by_text[key] = record
    return index


def lookup(index: CorpusIndex, gold: dict[str, Any]) -> dict[str, Any] | None:
    reference = gold.get("reference")
    if isinstance(reference, dict):
        original_id = reference.get("original_id")
        if isinstance(original_id, str) and original_id in index.by_id:
            return index.by_id[original_id]

    question_key = question_key_from_record(gold)
    if question_key and question_key in index.by_question:
        return index.by_question[question_key]

    for key in match_keys_from_record(gold):
        hit = index.by_text.get(key)
        if hit is not None:
            return hit
    return None


def pct(matches: int, total: int) -> float:
    if total == 0:
        return 0.0
    return round(100.0 * matches / total, 2)


def autosize_columns(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min(
            max(len(str(cell.value or "")) for cell in column_cells) + 2,
            max_width,
        )
        ws.column_dimensions[letter].width = width


def write_summary_sheet(
    wb: Workbook,
    *,
    corpus_name: str,
    total: int,
    topic_matches: int,
    name_matches: int,
    found: int,
) -> None:
    ws = wb.create_sheet(f"Summary_{corpus_name}")
    rows = [
        ("Corpus", corpus_name),
        ("Gold standard", str(DEFAULT_GOLD.relative_to(ROOT))),
        ("Total gold rows", total),
        ("Matched in corpus", found),
        ("Missing in corpus", total - found),
        ("Corpus coverage %", pct(found, total)),
        ("Topic label ID match % (of matched)", pct(topic_matches, found)),
        ("Topic label name match % (of matched)", pct(name_matches, found)),
        ("Topic label ID match % (of all gold)", pct(topic_matches, total)),
        ("Topic label name match % (of all gold)", pct(name_matches, total)),
        ("Topic label ID matches", topic_matches),
        ("Topic label name matches", name_matches),
    ]
    ws.append(["Metric", "Value"])
    for metric, value in rows:
        ws.append([metric, value])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    autosize_columns(ws)


def comparison_headers(corpus_label: str) -> list[str]:
    headers = [
        "id",
        "source_bucket",
        "source_file",
        "line_no",
        "question",
        "answers",
        "solution",
        "gold_topic_id",
        "gold_topic_name",
        "gold_reason",
    ]
    for tax in TAXONOMY_KEYS:
        headers.extend([
            f"gold_{tax}_id",
            f"gold_{tax}_name",
            f"gold_{tax}_reason",
        ])
    headers.extend([
        f"{corpus_label}_topic_id",
        f"{corpus_label}_topic_name",
        f"{corpus_label}_classifier_model",
        f"{corpus_label}_matched",
        f"{corpus_label}_topic_id_match",
        f"{corpus_label}_topic_name_match",
        "gold_notes",
    ])
    return headers


def build_comparison_row(
    gold: dict[str, Any],
    matched: dict[str, Any] | None,
    *,
    corpus_label: str,
) -> list[Any]:
    metadata = gold.get("metadata") or {}
    payload = gold.get("payload") or {}
    gold_cls = gold.get("classification") or {}

    gold_id, gold_name, gold_reason = classification_topic(gold_cls)

    row: list[Any] = [
        gold.get("id", ""),
        metadata.get("source_bucket", ""),
        metadata.get("source_file", ""),
        metadata.get("line_no", ""),
        payload.get("question") or gold.get("text", ""),
        format_answers(payload),
        format_solution(gold, payload),
        gold_id,
        gold_name,
        gold_reason,
    ]

    for tax in TAXONOMY_KEYS:
        tax_id, tax_name, tax_reason = taxonomy_field(gold_cls, tax)
        row.extend([tax_id, tax_name, tax_reason])

    if matched is None:
        row.extend(["", "", "", "no", "no", "no"])
    else:
        pred_cls = matched.get("classification") or {}
        pred_id, pred_name, _ = classification_topic(pred_cls)
        classifier = pred_cls.get("classifier") if isinstance(pred_cls, dict) else {}
        model = ""
        if isinstance(classifier, dict):
            model = str(classifier.get("model") or "")
        row.extend([
            pred_id,
            pred_name,
            model,
            "yes",
            "yes" if pred_id and pred_id == gold_id else "no",
            "yes" if pred_name and pred_name == gold_name else "no",
        ])

    row.append(str(gold_cls.get("notes") or ""))
    return row


def write_comparison_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    corpus_label: str,
    gold_rows: list[dict[str, Any]],
    index: CorpusIndex,
) -> tuple[int, int, int, int]:
    ws = wb.create_sheet(sheet_name)
    headers = comparison_headers(corpus_label)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    topic_matches = 0
    name_matches = 0
    found = 0

    for gold in gold_rows:
        matched = lookup(index, gold)
        row = build_comparison_row(gold, matched, corpus_label=corpus_label)
        ws.append(row)
        if matched is not None:
            found += 1
            gold_cls = gold.get("classification") or {}
            pred_cls = matched.get("classification") or {}
            gold_id, gold_name, _ = classification_topic(gold_cls)
            pred_id, pred_name, _ = classification_topic(pred_cls)
            if pred_id and pred_id == gold_id:
                topic_matches += 1
            if pred_name and pred_name == gold_name:
                name_matches += 1

    autosize_columns(ws)
    return topic_matches, name_matches, found, len(gold_rows)


def prediction_fields(matched: dict[str, Any] | None, *, prefix: str) -> dict[str, str]:
    if matched is None:
        return {
            f"{prefix}_topic_id": "",
            f"{prefix}_topic_name": "",
            f"{prefix}_model": "",
            f"{prefix}_matched": "no",
            f"{prefix}_topic_id_match": "no",
            f"{prefix}_topic_name_match": "no",
        }
    pred_cls = matched.get("classification") or {}
    pred_id, pred_name, _ = classification_topic(pred_cls)
    classifier = pred_cls.get("classifier") if isinstance(pred_cls, dict) else {}
    model = ""
    if isinstance(classifier, dict):
        model = str(classifier.get("model") or "")
    return {
        f"{prefix}_topic_id": pred_id,
        f"{prefix}_topic_name": pred_name,
        f"{prefix}_model": model,
        f"{prefix}_matched": "yes",
        f"{prefix}_topic_id_match": "",
        f"{prefix}_topic_name_match": "",
    }


def write_comparison_csv(
    *,
    output_path: Path,
    gold_rows: list[dict[str, Any]],
    embedding_index: CorpusIndex,
    nli_index: CorpusIndex,
    coarse: CoarseTaxonomy,
) -> dict[str, int]:
    fieldnames = [
        "id",
        "dataset",
        "source_file",
        "source_line_no",
        "question",
        "correct_answer",
        "gold_topic_id",
        "gold_topic_name",
        "gold_coarse_id",
        "gold_coarse_name",
        "gold_reason",
        "embedding_topic_id",
        "embedding_topic_name",
        "embedding_coarse_id",
        "embedding_coarse_name",
        "embedding_model",
        "embedding_matched",
        "embedding_topic_id_match",
        "embedding_topic_name_match",
        "embedding_coarse_id_match",
        "embedding_coarse_name_match",
        "nli_topic_id",
        "nli_topic_name",
        "nli_coarse_id",
        "nli_coarse_name",
        "nli_model",
        "nli_matched",
        "nli_topic_id_match",
        "nli_topic_name_match",
        "nli_coarse_id_match",
        "nli_coarse_name_match",
    ]

    stats = {
        "emb_topic_matches": 0,
        "emb_name_matches": 0,
        "emb_coarse_id_matches": 0,
        "emb_coarse_name_matches": 0,
        "emb_found": 0,
        "nli_topic_matches": 0,
        "nli_name_matches": 0,
        "nli_coarse_id_matches": 0,
        "nli_coarse_name_matches": 0,
        "nli_found": 0,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for gold in gold_rows:
            ref = gold.get("reference") or {}
            payload = gold.get("payload") or {}
            evaluation = gold.get("evaluation") or {}
            gold_cls = gold.get("classification") or {}
            gold_id, gold_name, gold_reason = classification_topic(gold_cls)
            gold_coarse_id, gold_coarse_name = coarse_group(gold_id, coarse)

            emb = lookup(embedding_index, gold)
            nli = lookup(nli_index, gold)

            row = {
                "id": gold.get("id", ""),
                "dataset": ref.get("dataset", ""),
                "source_file": ref.get("source_file", ""),
                "source_line_no": ref.get("source_line_no", ""),
                "question": payload.get("question", ""),
                "correct_answer": evaluation.get("correct_answer", ""),
                "gold_topic_id": gold_id,
                "gold_topic_name": gold_name,
                "gold_coarse_id": gold_coarse_id,
                "gold_coarse_name": gold_coarse_name,
                "gold_reason": gold_reason,
            }
            row.update(prediction_fields(emb, prefix="embedding"))
            row.update(prediction_fields(nli, prefix="nli"))
            row["embedding_coarse_id"] = ""
            row["embedding_coarse_name"] = ""
            row["embedding_coarse_id_match"] = "no"
            row["embedding_coarse_name_match"] = "no"
            row["nli_coarse_id"] = ""
            row["nli_coarse_name"] = ""
            row["nli_coarse_id_match"] = "no"
            row["nli_coarse_name_match"] = "no"

            if emb is not None:
                stats["emb_found"] += 1
                pred_id, pred_name, _ = classification_topic(emb.get("classification") or {})
                emb_coarse_id, emb_coarse_name = coarse_group(pred_id, coarse)
                row["embedding_coarse_id"] = emb_coarse_id
                row["embedding_coarse_name"] = emb_coarse_name
                row["embedding_topic_id_match"] = "yes" if pred_id and pred_id == gold_id else "no"
                row["embedding_topic_name_match"] = "yes" if pred_name and pred_name == gold_name else "no"
                row["embedding_coarse_id_match"] = (
                    "yes" if emb_coarse_id and emb_coarse_id == gold_coarse_id else "no"
                )
                row["embedding_coarse_name_match"] = (
                    "yes" if emb_coarse_name and emb_coarse_name == gold_coarse_name else "no"
                )
                if row["embedding_topic_id_match"] == "yes":
                    stats["emb_topic_matches"] += 1
                if row["embedding_topic_name_match"] == "yes":
                    stats["emb_name_matches"] += 1
                if row["embedding_coarse_id_match"] == "yes":
                    stats["emb_coarse_id_matches"] += 1
                if row["embedding_coarse_name_match"] == "yes":
                    stats["emb_coarse_name_matches"] += 1

            if nli is not None:
                stats["nli_found"] += 1
                pred_id, pred_name, _ = classification_topic(nli.get("classification") or {})
                nli_coarse_id, nli_coarse_name = coarse_group(pred_id, coarse)
                row["nli_coarse_id"] = nli_coarse_id
                row["nli_coarse_name"] = nli_coarse_name
                row["nli_topic_id_match"] = "yes" if pred_id and pred_id == gold_id else "no"
                row["nli_topic_name_match"] = "yes" if pred_name and pred_name == gold_name else "no"
                row["nli_coarse_id_match"] = (
                    "yes" if nli_coarse_id and nli_coarse_id == gold_coarse_id else "no"
                )
                row["nli_coarse_name_match"] = (
                    "yes" if nli_coarse_name and nli_coarse_name == gold_coarse_name else "no"
                )
                if row["nli_topic_id_match"] == "yes":
                    stats["nli_topic_matches"] += 1
                if row["nli_topic_name_match"] == "yes":
                    stats["nli_name_matches"] += 1
                if row["nli_coarse_id_match"] == "yes":
                    stats["nli_coarse_id_matches"] += 1
                if row["nli_coarse_name_match"] == "yes":
                    stats["nli_coarse_name_matches"] += 1

            writer.writerow(row)

    return stats


def print_agreement_summary(
    *,
    total: int,
    stats: dict[str, int],
    coarse_taxonomy_path: Path,
) -> None:
    emb_found = stats["emb_found"]
    nli_found = stats["nli_found"]
    print(f"Coarse taxonomy: {coarse_taxonomy_path.relative_to(ROOT)}")
    print(
        f"Fine   — Embedding: {stats['emb_topic_matches']}/{emb_found} "
        f"({pct(stats['emb_topic_matches'], emb_found)}% topic ID), "
        f"NLI: {stats['nli_topic_matches']}/{nli_found} "
        f"({pct(stats['nli_topic_matches'], nli_found)}% topic ID)"
    )
    print(
        f"Coarse — Embedding: {stats['emb_coarse_id_matches']}/{emb_found} "
        f"({pct(stats['emb_coarse_id_matches'], emb_found)}% group ID), "
        f"NLI: {stats['nli_coarse_id_matches']}/{nli_found} "
        f"({pct(stats['nli_coarse_id_matches'], nli_found)}% group ID)"
    )
    print(f"Rows in gold standard: {total}")


def export_xlsx(
    *,
    gold_path: Path,
    output_path: Path,
    output_csv: Path | None = None,
    coarse_taxonomy_path: Path = DEFAULT_COARSE_TAXONOMY,
) -> None:
    gold_rows = list(iter_jsonl(gold_path))
    coarse = load_coarse_taxonomy(coarse_taxonomy_path)

    embedding_index = build_index(
        [PROCESSED],
        exclude_dirs=EMBEDDING_EXCLUDE_DIRS,
    )
    nli_index = build_index(list(NLI_ROOTS))

    csv_stats: dict[str, int] | None = None
    if output_csv is not None:
        csv_stats = write_comparison_csv(
            output_path=output_csv,
            gold_rows=gold_rows,
            embedding_index=embedding_index,
            nli_index=nli_index,
            coarse=coarse,
        )
        print(f"Wrote {output_csv}")

    wb = Workbook()
    wb.remove(wb.active)

    emb_topic, emb_name, emb_found, total = write_comparison_sheet(
        wb,
        sheet_name="Embedding_Comparison",
        corpus_label="embedding",
        gold_rows=gold_rows,
        index=embedding_index,
    )
    write_summary_sheet(
        wb,
        corpus_name="Embedding",
        total=total,
        topic_matches=emb_topic,
        name_matches=emb_name,
        found=emb_found,
    )

    nli_topic, nli_name, nli_found, _ = write_comparison_sheet(
        wb,
        sheet_name="NLI_Comparison",
        corpus_label="nli",
        gold_rows=gold_rows,
        index=nli_index,
    )
    write_summary_sheet(
        wb,
        corpus_name="NLI",
        total=total,
        topic_matches=nli_topic,
        name_matches=nli_name,
        found=nli_found,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote {output_path}")
    total = len(gold_rows)
    if csv_stats is not None:
        print_agreement_summary(
            total=total,
            stats=csv_stats,
            coarse_taxonomy_path=coarse_taxonomy_path,
        )
    print(
        f"XLSX fine — Embedding: {emb_found}/{total} matched, "
        f"{pct(emb_topic, emb_found)}% topic ID agreement"
    )
    print(
        f"XLSX fine — NLI: {nli_found}/{total} matched, "
        f"{pct(nli_topic, nli_found)}% topic ID agreement"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gold", type=Path, default=DEFAULT_GOLD)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=DEFAULT_OUTPUT_CSV,
        help="Combined embedding vs NLI comparison CSV (pass empty string to skip).",
    )
    parser.add_argument(
        "--coarse-taxonomy",
        type=Path,
        default=DEFAULT_COARSE_TAXONOMY,
        help="Coarse group taxonomy JSON (default: schema/taxonomy_coarse.json).",
    )
    args = parser.parse_args()
    csv_path = args.output_csv if str(args.output_csv) else None
    export_xlsx(
        gold_path=args.gold,
        output_path=args.output,
        output_csv=csv_path,
        coarse_taxonomy_path=args.coarse_taxonomy,
    )


if __name__ == "__main__":
    main()
